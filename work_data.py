import os
import time
import sqlite3
import openpyxl as ex
import shutil
from zipfile import ZipFile


def get_list_files():
    list_files = []
    for filename in os.listdir("input"):
        list_files.append("input/"+filename)
    return list_files


def process(files, stores):
    header = ["Код отдела", "Наименование отдела", "Операция",
              "Код операции", "Номер движ", "Дата движения", "Поставщик",
              "ИНН Поставщика", "Column7", "Column8", "Column9", "Column10",
              "Column11", "Column12", "Column13", "Column14", "кОД товара",
              "Наименование товара", "код производителя", "производитель", "страна производителя",
              "ШК", "Цена за единицу закупочная в рублях, округление до целых копеек",
              "Цена за единицу розничная в рублях, округление до целых копеек.",
              "Количество", "Серия товара", "Срок годности товара",
              "Уникальный код партии из системы учёта товародвижения, в формате штрих кода EAN-13"]
    tail = str(time.time())
    fw_move = open(f"move {tail}.csv", 'w')
    fw_move.write(";".join(map(str, header)))
    fw_move.write("\n")
    fw_base = open(f"base {tail}.csv", 'w')
    fw_base.write(";".join(map(str, header)))
    fw_base.write("\n")
    list_stores_fact = {'.ost': [], '.mov': []}
    for filename in files:
        file_extension = os.path.splitext(filename)
        if file_extension[1] != '.mov' and file_extension[1] != '.ost':
            continue
        with open(filename, 'r') as fr:
            code_store = filename[0:filename.find("$")].replace("input/", "")
            list_stores_fact[file_extension[1]].append(code_store)
            if code_store in stores:
                name_store = stores[code_store]
            else:
                name_store = ""
            for line in fr:
                line = line.split(";")
                if line[0] == '0':
                    oper = 'остаток'
                elif line[0] == '10':
                    oper = 'приход от поставщика'
                elif line[0] == '19':
                    oper = 'другие виды прихода'
                elif line[0] == '20':
                    oper = 'продажа товара через кассу конечному потребителю'
                elif line[0] == '29':
                    oper = 'другие виды расхода'
                else:
                    oper = ""
                if line[0] == '0':
                    fw_base.write(code_store + ";")
                    fw_base.write(name_store + ";")
                    fw_base.write(oper + ";")
                    line[-1] = line[-1].split()
                    fw_base.write(";".join(map(str, line)))
                    fw_base.write("\n")
                else:
                    fw_move.write(code_store+";")
                    fw_move.write(name_store+";")
                    fw_move.write(oper+";")
                    line[-1] = line[-1].split()
                    fw_move.write(";".join(map(str, line)))
                    fw_move.write("\n")
    fw_move.close()
    fw_base.close()
    print(f"Обработка завершена\n")
    return list_stores_fact


def get_stores():
    stores = {}
    conn = sqlite3.connect('settings.db')
    cursor = conn.cursor()
    cursor.execute('select * from stores')
    rows = cursor.fetchall()
    for row in rows:
        stores[row[0]] = row[1]
    cursor.close()
    conn.close()
    return stores


def check_avail_stores(list_stores_fact, stores):
    if convert_xlsx("X:\\\\АптекиМФО\\СписокАптек.xlsx", "list_pharmacy.xlsx") is False:
        return False
    conn = sqlite3.connect('settings.db')
    cursor = conn.cursor()
    codes = []
    rows = cursor.execute('select * from code')
    for row in rows:
        codes.append(row[0])
    wb = ex.load_workbook("list_pharmacy.xlsx")
    sheet = wb.active
    last_row = sheet.max_row + 1
    for i in range(5, last_row):
        brand_code = sheet.cell(row=i, column=7).value
        fl_del = sheet.cell(row=i, column=14).value
        virt_skl = sheet.cell(row=i, column=15).value
        code1c = sheet.cell(row=i, column=3).value
        partner1c = sheet.cell(row=i, column=2).value
        if brand_code is None or fl_del == 'Да' or virt_skl == 'Да':
            continue
        if brand_code not in codes:
            continue
        cursor.execute(f"select * from replace where code1c='{code1c}'")
        res = cursor.fetchall()
        if len(res) == 0:
            print(f"Отсутствует аптека в выгрузке {partner1c} код {code1c}")
        else:
            for row in res:
                code_ = row[2]
                name_ = row[3]
                if code_ not in list_stores_fact['.mov']:
                    print(f"Отсутствует файл движение по аптеке {name_} с кодом {code_}")
                if code_ not in list_stores_fact['.ost']:
                    print(f"Отсутствует файл остатков по аптеке {name_} с кодом {code_}")
    cursor.close()
    conn.close()
    if os.path.exists("list_pharmacy.xlsx"):
        os.remove("list_pharmacy.xlsx")


def convert_xlsx(file_in, new_filename):
    if os.path.exists(file_in) is False:
        print(f"Файл {file_in} не найден")
        return False
    tmp_folder = '/tmp/convert_wrong_excel/'
    os.makedirs(tmp_folder, exist_ok=True)
    with ZipFile(file_in) as excel_container:
        excel_container.extractall(tmp_folder)
    wrong_file_path = os.path.join(tmp_folder, 'xl', 'SharedStrings.xml')
    correct_file_path = os.path.join(tmp_folder, 'xl', 'sharedStrings.xml')
    os.rename(wrong_file_path, correct_file_path)
    shutil.make_archive('tmp', 'zip', tmp_folder)
    os.replace('tmp.zip', new_filename)
    return True


